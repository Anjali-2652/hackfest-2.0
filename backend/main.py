"""
PolicyGuard Compliance Backend - FastAPI Server
Serves compliance data from the PolicyGuard Excel dataset
"""

import os
import sys
from pathlib import Path
from datetime import datetime

# Ensure we're in the right directory
project_root = Path(__file__).parent.parent
os.chdir(project_root)
sys.path.insert(0, str(project_root))

from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import json
import openpyxl

app = FastAPI(
    title="PolicyGuard API",
    version="1.0.0",
    description="Compliance Dashboard API for PolicyGuard"
)

# Enable CORS for frontend communication
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Ahmed Tariq (Sales) - HIGH RISK - Client-facing, aggressive sales tactics, many violations
# Violations: Unauthorized discounts, client communication breaches, credit policy violations
MONTHLY_DATA_EMP001 = [
    {"month": "January", "compliant": 45, "violation": 78},
    {"month": "February", "compliant": 52, "violation": 71},
    {"month": "March", "compliant": 48, "violation": 75},
    {"month": "April", "compliant": 56, "violation": 67},
    {"month": "May", "compliant": 50, "violation": 73},
    {"month": "June", "compliant": 59, "violation": 64},
    {"month": "July", "compliant": 61, "violation": 62},
    {"month": "August", "compliant": 58, "violation": 65},
    {"month": "September", "compliant": 63, "violation": 60},
    {"month": "October", "compliant": 60, "violation": 63},
    {"month": "November", "compliant": 65, "violation": 58},
    {"month": "December", "compliant": 62, "violation": 61}
]

# Sarah Mitchell (Operations) - LOW RISK - Excellent process compliance
# Violations: Minor documentation gaps, scheduling adjustments, procedure updates
MONTHLY_DATA_EMP002 = [
    {"month": "January", "compliant": 218, "violation": 5},
    {"month": "February", "compliant": 224, "violation": 3},
    {"month": "March", "compliant": 221, "violation": 4},
    {"month": "April", "compliant": 226, "violation": 2},
    {"month": "May", "compliant": 229, "violation": 1},
    {"month": "June", "compliant": 227, "violation": 2},
    {"month": "July", "compliant": 231, "violation": 1},
    {"month": "August", "compliant": 225, "violation": 3},
    {"month": "September", "compliant": 223, "violation": 4},
    {"month": "October", "compliant": 228, "violation": 2},
    {"month": "November", "compliant": 230, "violation": 1},
    {"month": "December", "compliant": 232, "violation": 0}
]

# James Chen (IT) - MEDIUM RISK - Security focused, improving over time
# Violations: Unauthorized system access, weak password policies, unpatched systems
MONTHLY_DATA_EMP003 = [
    {"month": "January", "compliant": 32, "violation": 36},
    {"month": "February", "compliant": 35, "violation": 33},
    {"month": "March", "compliant": 38, "violation": 30},
    {"month": "April", "compliant": 40, "violation": 28},
    {"month": "May", "compliant": 42, "violation": 26},
    {"month": "June", "compliant": 45, "violation": 23},
    {"month": "July", "compliant": 48, "violation": 20},
    {"month": "August", "compliant": 50, "violation": 18},
    {"month": "September", "compliant": 52, "violation": 16},
    {"month": "October", "compliant": 54, "violation": 14},
    {"month": "November", "compliant": 56, "violation": 12},
    {"month": "December", "compliant": 58, "violation": 10}
]

# Simulate database storage
DATA_STORAGE = {
    "records": [],
    "metadata": {
        "total_records": 4000,
        "compliant": 604,
        "violations": 3396,
        "compliance_rate": "15.1%",
        "last_updated": datetime.now().isoformat()
    },
    "employees": {
        "EMP-001": {
            "name": "Ahmed Tariq",
            "department": "Sales",
            "avatar": "AT",
            "avatarColor": "#FF8A65",
            "violations": 27,
            "total_records": 45,
            "violation_rate": 60,
            "css_avg": 3.2,
            "orgAvgViolationRate": 86.9,
            "policyMinCss": 3.5,
            "targets_hit": 4,
            "total_targets": 12,
            "risk_level": "HIGH",
            "monthlyData": MONTHLY_DATA_EMP001,
            "violation_types": ["Unauthorized Discounts", "Client Communication Breach", "Credit Policy Violation", "Sales Practice Violation"]
        },
        "EMP-002": {
            "name": "Sarah Mitchell",
            "department": "Operations",
            "avatar": "SM",
            "avatarColor": "#42A5F5",
            "violations": 2,
            "total_records": 40,
            "violation_rate": 5,
            "css_avg": 9.1,
            "orgAvgViolationRate": 86.9,
            "policyMinCss": 3.5,
            "targets_hit": 12,
            "total_targets": 12,
            "risk_level": "LOW",
            "monthlyData": MONTHLY_DATA_EMP002,
            "violation_types": ["Documentation Gap", "Procedure Update", "Scheduling Adjustment"]
        },
        "EMP-003": {
            "name": "James Chen",
            "department": "IT",
            "avatar": "JC",
            "avatarColor": "#66BB6A",
            "violations": 16,
            "total_records": 34,
            "violation_rate": 47,
            "css_avg": 6.8,
            "orgAvgViolationRate": 86.9,
            "policyMinCss": 3.5,
            "targets_hit": 8,
            "total_targets": 12,
            "risk_level": "MEDIUM",
            "monthlyData": MONTHLY_DATA_EMP003,
            "violation_types": ["Unauthorized System Access", "Weak Password Policy", "Unpatched System", "Security Protocol Breach"]
        }
    }
}

def extract_excel_data(file_path: str):
    """Extract data from PolicyGuard Excel file"""
    try:
        if not Path(file_path).exists():
            print(f"⚠ Excel file not found: {file_path}")
            return None
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Extract summary data from row 4
        total_records = ws['B4'].value or 4000
        compliant = ws['D4'].value or 604
        violations = ws['F4'].value or 3396
        compliance_rate = ws['H4'].value or "15.1%"
        
        return {
            "summary": {
                "total_records": int(total_records),
                "compliant": int(compliant),
                "violations": int(violations),
                "compliance_rate": str(compliance_rate)
            }
        }
    except Exception as e:
        print(f"⚠ Error extracting Excel data: {e}")
        return None

# Try to load initial data from Excel file
excel_path = r"c:\Users\Acer Nitro 5\Downloads\PolicyGuard_Compliance_Report.xlsx"
initial_data = extract_excel_data(excel_path)
if initial_data:
    DATA_STORAGE["metadata"].update(initial_data["summary"])
    print(f"✓ Loaded PolicyGuard data: {initial_data['summary']}")
else:
    print(f"✓ Using default data (Excel file not found)")

@app.get("/")
def root():
    return {"message": "PolicyGuard API", "status": "running"}

@app.get("/api/dashboard/summary")
def get_dashboard_summary():
    """Get compliance dashboard summary"""
    return {
        "total_records": DATA_STORAGE["metadata"]["total_records"],
        "compliant": DATA_STORAGE["metadata"]["compliant"],
        "violations": DATA_STORAGE["metadata"]["violations"],
        "compliance_rate": DATA_STORAGE["metadata"]["compliance_rate"],
        "violation_risk": "HIGH" if DATA_STORAGE["metadata"]["violations"] > DATA_STORAGE["metadata"]["compliant"] else "MEDIUM"
    }

@app.get("/api/dashboard/monthly-violations")
def get_monthly_violations(employee_id: str = "EMP-001"):
    """Get monthly violation trends for an employee"""
    emp = DATA_STORAGE["employees"].get(employee_id)
    if emp and "monthlyData" in emp:
        return emp["monthlyData"]
    return MONTHLY_DATA_EMP001

@app.get("/api/dashboard/violation-reasons")
def get_violation_reasons():
    """Get breakdown of violation reasons"""
    return [
        {
            "reason": "Low Customer Satisfaction",
            "count": 2614,
            "percentage": 76.9,
            "severity": "HIGH"
        },
        {
            "reason": "Policy Exception",
            "count": 431,
            "percentage": 12.7,
            "severity": "MEDIUM"
        },
        {
            "reason": "Training Gap",
            "count": 351,
            "percentage": 10.3,
            "severity": "MEDIUM"
        }
    ]

@app.get("/api/employees/{employee_id}")
def get_employee_data(employee_id: str):
    """Get employee-specific compliance data"""
    if employee_id in DATA_STORAGE["employees"]:
        emp = DATA_STORAGE["employees"][employee_id]
        return {
            "id": employee_id,
            "name": emp.get("name"),
            "avatar": emp.get("avatar"),
            "avatarColor": emp.get("avatarColor"),
            "department": emp.get("department"),
            "violations": emp.get("violations"),
            "total_records": emp.get("total_records"),
            "violation_rate": emp.get("violation_rate"),
            "css_avg": emp.get("css_avg"),
            "orgAvgViolationRate": emp.get("orgAvgViolationRate"),
            "policyMinCss": emp.get("policyMinCss"),
            "targets_hit": emp.get("targets_hit"),
            "total_targets": emp.get("total_targets"),
            "risk_level": emp.get("risk_level"),
            "monthlyData": emp.get("monthlyData", MONTHLY_DATA_EMP001)
        }
    
    # Return default for unknown employees
    return {
        "id": employee_id,
        "name": "Unknown Employee",
        "violations": 0,
        "total_records": 0,
        "violation_rate": 0,
        "css_avg": 0,
        "targets_hit": 0,
        "total_targets": 0,
        "risk_level": "LOW",
        "monthlyData": MONTHLY_DATA_EMP001
    }

@app.post("/api/records/upload")
async def upload_records(file: UploadFile = File(...)):
    """Upload and process compliance records"""
    try:
        # Save the file temporarily
        temp_path = f"temp_{file.filename}"
        with open(temp_path, "wb") as f:
            content = await file.read()
            f.write(content)
        
        # Extract data from the file
        df = pd.read_excel(temp_path)
        
        # Store records
        DATA_STORAGE["records"] = df.to_dict('records')
        
        # Update metadata
        DATA_STORAGE["metadata"]["total_records"] = len(df)
        
        # Clean up temp file
        Path(temp_path).unlink()
        
        return {
            "status": "success",
            "message": f"Uploaded {len(df)} records",
            "total_records": len(df)
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/records")
def get_records(skip: int = 0, limit: int = 50):
    """Get compliance records with pagination"""
    records = DATA_STORAGE["records"][skip:skip + limit]
    return {
        "records": records,
        "total": len(DATA_STORAGE["records"]),
        "skip": skip,
        "limit": limit
    }

@app.get("/api/records/stats")
def get_records_stats():
    """Get statistics from records"""
    if not DATA_STORAGE["records"]:
        return {
            "total": 0,
            "compliant": 0,
            "violations": 0,
            "compliance_rate": 0
        }
    
    total = len(DATA_STORAGE["records"])
    compliant = sum(1 for r in DATA_STORAGE["records"] if r.get("Status") == "Compliant")
    violations = total - compliant
    
    return {
        "total": total,
        "compliant": compliant,
        "violations": violations,
        "compliance_rate": round((compliant / total * 100), 1) if total > 0 else 0
    }

@app.get("/api/health")
def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "version": "1.0.0"}

if __name__ == "__main__":
    import uvicorn
    print("\n" + "="*60)
    print("🛡  PolicyGuard API Server")
    print("="*60)
    print(f"Starting FastAPI server...")
    print(f"📍 API will be available at: http://127.0.0.1:8000")
    print(f"📚 API Docs at: http://127.0.0.1:8000/docs")
    print(f"⚙️  ReDoc at: http://127.0.0.1:8000/redoc")
    print("="*60 + "\n")
    
    uvicorn.run(app, host="127.0.0.1", port=8000, log_level="info")
